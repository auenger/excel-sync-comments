import openpyxl
from openpyxl.comments import Comment
from openpyxl.utils import column_index_from_string


# ================= 配置加载 =================
import os
import sys
import configparser

# 获取可执行文件所在目录（支持 PyInstaller 打包）
if getattr(sys, 'frozen', False):
    # 如果是打包后的可执行文件
    SCRIPT_DIR = os.path.dirname(sys.executable)
else:
    # 如果是源代码运行
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

CONFIG_FILE = os.path.join(SCRIPT_DIR, 'config.ini')

# 默认配置（如果配置文件不存在时使用）
DEFAULT_CONFIG = {
    'SOURCE_FILE': 'source.xlsx',
    'TARGET_FILE': 'target.xlsx',
    'OUTPUT_FILE': 'target_updated.xlsx',
    'COL_REGION': 'C',
    'COL_NAME': 'B',
    'TARGET_REGION': ['厦门', '广州', '长泰', '龙岩', '南昌', '石家庄'],
    'COLS_TO_SYNC': ['DO', 'DP', 'DS', 'DU'],
    'MERGE_COMMENTS': True,
    'MERGE_SEPARATOR': '\n---\n',
    'START_ROW': 3
}

def load_config():
    """从配置文件加载配置，如果文件不存在则使用默认配置"""
    config = DEFAULT_CONFIG.copy()
    
    if not os.path.exists(CONFIG_FILE):
        print(f"警告: 配置文件 {CONFIG_FILE} 不存在，使用默认配置")
        return config
    
    try:
        parser = configparser.ConfigParser()
        parser.read(CONFIG_FILE, encoding='utf-8')
        
        # 读取文件路径
        if parser.has_section('文件路径'):
            config['SOURCE_FILE'] = parser.get('文件路径', '源文件', fallback=config['SOURCE_FILE'])
            config['TARGET_FILE'] = parser.get('文件路径', '目标文件', fallback=config['TARGET_FILE'])
            config['OUTPUT_FILE'] = parser.get('文件路径', '输出文件', fallback=config['OUTPUT_FILE'])
        
        # 读取列配置
        if parser.has_section('列配置'):
            config['COL_REGION'] = parser.get('列配置', '区域列', fallback=config['COL_REGION']).strip()
            config['COL_NAME'] = parser.get('列配置', '姓名列', fallback=config['COL_NAME']).strip()
            
            # 读取同步列（逗号分隔）
            cols_str = parser.get('列配置', '同步列', fallback='')
            if cols_str:
                config['COLS_TO_SYNC'] = [col.strip() for col in cols_str.split(',') if col.strip()]
        
        # 读取筛选条件
        if parser.has_section('筛选条件'):
            region_str = parser.get('筛选条件', '筛选区域', fallback='').strip()
            if region_str and region_str.lower() != 'none':
                # 多个区域用逗号分隔
                config['TARGET_REGION'] = [r.strip() for r in region_str.split(',') if r.strip()]
            else:
                config['TARGET_REGION'] = None
        
        # 读取数据设置
        if parser.has_section('数据设置'):
            config['START_ROW'] = parser.getint('数据设置', '数据起始行', fallback=config['START_ROW'])
        
        # 读取批注合并设置
        if parser.has_section('批注合并'):
            config['MERGE_COMMENTS'] = parser.getboolean('批注合并', '启用合并', fallback=config['MERGE_COMMENTS'])
            separator = parser.get('批注合并', '分隔符', fallback=config['MERGE_SEPARATOR'])
            # 处理转义字符
            config['MERGE_SEPARATOR'] = separator.replace('\\n', '\n').replace('\\t', '\t')
        
        print(f"✓ 已从配置文件加载配置: {CONFIG_FILE}\n")
        
    except Exception as e:
        print(f"警告: 读取配置文件时出错: {e}")
        print("将使用默认配置\n")
    
    return config

# 加载配置
config = load_config()

# 将配置赋值给变量（保持原有代码兼容）
SOURCE_FILE = os.path.join(SCRIPT_DIR, config['SOURCE_FILE'])
TARGET_FILE = os.path.join(SCRIPT_DIR, config['TARGET_FILE'])
OUTPUT_FILE = os.path.join(SCRIPT_DIR, config['OUTPUT_FILE'])
COL_REGION = config['COL_REGION']
COL_NAME = config['COL_NAME']
TARGET_REGION = config['TARGET_REGION']
COLS_TO_SYNC = config['COLS_TO_SYNC']
MERGE_COMMENTS = config['MERGE_COMMENTS']
MERGE_SEPARATOR = config['MERGE_SEPARATOR']
START_ROW = config['START_ROW']
# =======================================================


def sync_excel_comments():
    print(f"正在加载源文件: {SOURCE_FILE} ...")
    wb_source = openpyxl.load_workbook(SOURCE_FILE, data_only=False)
    ws_source = wb_source.active # 或者指定 sheet 名: wb_source['Sheet1']

    print(f"正在加载目标文件: {TARGET_FILE} ...")
    wb_target = openpyxl.load_workbook(TARGET_FILE, data_only=False)
    ws_target = wb_target.active

    # --- 第一步：构建源数据的批注映射表 ---
    # 结构: { "张三": { "DN": "批注内容A", "DO": "批注内容B" } }
    comments_map = {}
    
    print("正在建立源数据索引...")
    # 获取列的数字索引 (openpyxl 内部用)
    idx_region = column_index_from_string(COL_REGION) - 1
    idx_name = column_index_from_string(COL_NAME) - 1
    
    for row in ws_source.iter_rows(min_row=START_ROW):
        # 1. 区域筛选 - 支持多区域
        region_val = row[idx_region].value
        
        # 如果设置了区域筛选，且当前行不匹配，则跳过
        if TARGET_REGION:
            # 将 TARGET_REGION 统一转换为列表处理
            regions = TARGET_REGION if isinstance(TARGET_REGION, list) else [TARGET_REGION]
            if region_val not in regions:
                continue
            
        # 2. 获取姓名
        name_val = row[idx_name].value
        if not name_val:
            continue
            
        # 3. 提取指定列的批注
        comments_map[name_val] = {}
        for col_letter in COLS_TO_SYNC:
            # 拼接单元格坐标，例如 "DN5"
            cell_ref = f"{col_letter}{row[0].row}"
            source_cell = ws_source[cell_ref]
            
            if source_cell.comment:
                # 存下批注对象
                comments_map[name_val][col_letter] = source_cell.comment

    print(f"索引完成，共找到 {len(comments_map)} 个符合区域筛选的人员数据。")
    if TARGET_REGION:
        regions = TARGET_REGION if isinstance(TARGET_REGION, list) else [TARGET_REGION]
        print(f"筛选区域: {', '.join(regions)}")

    # --- 第二步：写入目标文件 ---
    print("正在同步批注到目标文件...")
    updated_count = 0
    merged_count = 0
    
    # 记录详细信息
    sync_details = []  # 记录所有同步操作的详情
    merged_details = []  # 记录合并操作的详情
    
    for row in ws_target.iter_rows(min_row=START_ROW):
        name_val = row[idx_name].value
        
        # 如果这个人在我们的源映射表中
        if name_val in comments_map:
            user_comments = comments_map[name_val]
            
            for col_letter, source_comment in user_comments.items():
                # 定位目标单元格
                target_cell_ref = f"{col_letter}{row[0].row}"
                target_cell = ws_target[target_cell_ref]
                
                # 检查目标单元格是否已有批注
                if target_cell.comment and MERGE_COMMENTS:
                    # 合并批注：原有批注 + 分隔符 + 新批注
                    existing_text = target_cell.comment.text
                    new_text = existing_text + MERGE_SEPARATOR + source_comment.text
                    
                    # 保留原有批注的作者，或者使用源批注的作者
                    author = target_cell.comment.author or source_comment.author
                    new_comment = Comment(new_text, author)
                    merged_count += 1
                    
                    # 记录合并详情
                    merged_details.append({
                        'name': name_val,
                        'cell': target_cell_ref,
                        'column': col_letter,
                        'original': existing_text[:50] + '...' if len(existing_text) > 50 else existing_text,
                        'new': source_comment.text[:50] + '...' if len(source_comment.text) > 50 else source_comment.text
                    })
                    
                    sync_details.append(f"[合并] {name_val} - 列{col_letter} (单元格{target_cell_ref})")
                else:
                    # 直接创建新批注（覆盖或新增）
                    new_comment = Comment(source_comment.text, source_comment.author)
                    action = "覆盖" if target_cell.comment else "新增"
                    sync_details.append(f"[{action}] {name_val} - 列{col_letter} (单元格{target_cell_ref})")
                
                # 写入批注
                target_cell.comment = new_comment
                updated_count += 1

    # --- 第三步：保存 ---
    wb_target.save(OUTPUT_FILE)
    
    # --- 第四步：生成日志文件 ---
    from datetime import datetime
    log_file = os.path.join(SCRIPT_DIR, f'sync_log_{datetime.now().strftime("%Y%m%d_%H%M%S")}.txt')
    
    with open(log_file, 'w', encoding='utf-8') as f:
        f.write("=" * 80 + "\n")
        f.write("Excel 批注同步日志\n")
        f.write("=" * 80 + "\n\n")
        
        f.write(f"执行时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        
        # 配置信息
        f.write("【配置信息】\n")
        f.write(f"源文件: {SOURCE_FILE}\n")
        f.write(f"目标文件: {TARGET_FILE}\n")
        f.write(f"输出文件: {OUTPUT_FILE}\n")
        f.write(f"区域列: {COL_REGION}\n")
        f.write(f"姓名列: {COL_NAME}\n")
        
        if TARGET_REGION:
            regions = TARGET_REGION if isinstance(TARGET_REGION, list) else [TARGET_REGION]
            f.write(f"筛选区域: {', '.join(regions)}\n")
        else:
            f.write("筛选区域: 无筛选\n")
            
        f.write(f"同步列: {', '.join(COLS_TO_SYNC)}\n")
        f.write(f"批注合并: {'启用' if MERGE_COMMENTS else '禁用'}\n")
        f.write(f"数据起始行: {START_ROW}\n\n")
        
        # 统计信息
        f.write("【统计信息】\n")
        f.write(f"匹配人员数: {len(comments_map)} 人\n")
        f.write(f"同步批注数: {updated_count} 个\n")
        f.write(f"合并批注数: {merged_count} 个\n")
        f.write(f"新增/覆盖数: {updated_count - merged_count} 个\n\n")
        
        # 详细操作记录
        f.write("=" * 80 + "\n")
        f.write("【详细操作记录】\n")
        f.write("=" * 80 + "\n\n")
        
        for detail in sync_details:
            f.write(detail + "\n")
        
        # 合并详情
        if merged_details:
            f.write("\n" + "=" * 80 + "\n")
            f.write("【合并批注详情】\n")
            f.write("=" * 80 + "\n\n")
            
            for idx, item in enumerate(merged_details, 1):
                f.write(f"{idx}. 姓名: {item['name']} | 列: {item['column']} | 单元格: {item['cell']}\n")
                f.write(f"   原批注: {item['original']}\n")
                f.write(f"   新批注: {item['new']}\n")
                f.write("\n")
    
    # 控制台输出
    print(f"处理完成！成功同步了 {updated_count} 个批注。")
    if merged_count > 0:
        print(f"其中 {merged_count} 个批注与原有批注进行了合并。")
    print(f"文件已保存为: {OUTPUT_FILE}")
    print(f"详细日志已保存为: {log_file}")

if __name__ == '__main__':
    sync_excel_comments()